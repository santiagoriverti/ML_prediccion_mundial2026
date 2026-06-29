# -*- coding: utf-8 -*-
"""
enriquecer_excel.py
===================
Script REPRODUCIBLE que enriquece ``Mundial_2026_fuente_datos.xlsx`` con las
variables predictoras curadas y las formulas de auto-actualizacion del Excel.

Por que un script y no edicion manual:
  * Deja TRAZA y AUDITORIA de cada dato curado (con su fuente/metodo).
  * Es re-ejecutable: si se corrige un dato, se regenera el Excel identico.

Que escribe (idempotente):
  1. Hoja **DTs**: puntaje de trayectoria del DT (seleccion 0-50 + clubes/ligas
     0-50 = total 0-100). VALORES LITERALES (el modelo los lee directo).
  2. Hoja **Clasificatorias**: registro de la eliminatoria 2026 (PJ/PG/PE/PP/GF/GC)
     + dificultad de la confederacion (segun mundiales que produce) + puntaje
     ponderado = %Pts * dificultad. LITERALES.
  3. Hoja **Predictores_pais**: "Jug. en top-5 ligas" (conteo sobre plantel de 26).
     LITERAL.
  4. Hoja **Eliminatorias**: los slots 1o/2o (p.ej. "1o C") se resuelven por
     FORMULA contra la hoja Posiciones (se auto-actualizan al cargar grupos);
     los slots de "mejor tercero" se dejan al motor Python (fuente de verdad).
  5. Hoja **Leeme**: nota de corte actualizada.

IMPORTANTE sobre datos curados: las cifras de DTs, Clasificatorias y % top-5 son
ESTIMACIONES razonables al estado ~early-2026 (conocimiento experto + prensa
deportiva), NO cifras oficiales exactas. El objetivo es aportar SENAL ordinal
(quien tiene mejor DT / mejor clasificatoria / mas jugadores de elite), no
precision contable. El modelo usa diferencias A-B, robustas a errores chicos.

Uso:  python scripts/enriquecer_excel.py
"""
from __future__ import annotations

import os
import openpyxl

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(RAIZ, "Mundial_2026_fuente_datos.xlsx")

# ===========================================================================
# 1) PUNTAJE DE DIRECTORES TECNICOS  (seleccion 0-50, clubes/ligas 0-50)
# ---------------------------------------------------------------------------
# Rubrica:
#   * Seleccion: titulos mundiales/continentales como DT, profundidad en
#     mundiales, experiencia internacional y prestigio (0-50).
#   * Clubes/ligas: titulos de liga top-5 / Champions, jerarquia de los clubes
#     dirigidos y experiencia en ligas top (0-50).
# DT_SCORE[pais] = (sel, club)
# ===========================================================================
DT_SCORE = {
    "México": (24, 22),            # Aguirre: Mexico x3, Japon; larga carrera en LaLiga
    "Sudáfrica": (30, 10),         # Broos: campeon AFCON 2017 con Camerun
    "Corea del Sur": (18, 12),     # Hong Myung-bo: bronce olimpico 2012; K-League
    "Chequia": (8, 12),            # Koubek: liga checa
    "Canadá": (14, 22),            # Marsch: Leipzig/Salzburg/Leeds
    "Bosnia y Herzegovina": (6, 6),# Barbarez: poca experiencia de DT
    "Catar": (18, 30),            # Lopetegui: Sevilla (Europa League), Real Madrid, PL
    "Suiza": (20, 16),             # Yakin: WC2022 R16; titulos con Basilea
    "Brasil": (8, 50),             # Ancelotti: 5 Champions, ligas en los 5 grandes
    "Marruecos": (12, 8),          # Ouahbi: exito sub-20; nuevo en la mayor
    "Haití": (12, 6),              # Migne: experiencia AFCON (Kenia, G. Ecuatorial)
    "Escocia": (22, 14),           # Clarke: 2 Euros con Escocia; asist. top clubs
    "Estados Unidos": (12, 36),    # Pochettino: final Champions con Tottenham, PSG
    "Paraguay": (24, 20),          # Alfaro: Ecuador WC2022, Costa Rica; Boca
    "Australia": (16, 16),         # Popovic: ACL con Western Sydney
    "Turquía": (16, 22),           # Montella: Milan, Fiorentina, Sevilla
    "Alemania": (22, 30),          # Nagelsmann: Bayern, Leipzig
    "Curazao": (30, 26),           # Advocaat: Holanda/Corea/Rusia; Rangers, Zenit
    "Costa de Marfil": (26, 6),    # Faé: campeon AFCON 2023
    "Ecuador": (16, 16),           # Beccacece: clubes argentinos, Elche
    "Países Bajos": (26, 30),      # Koeman: Barcelona, Ajax; final Nations League
    "Japón": (24, 14),             # Moriyasu: WC2022 (gano a Alemania y Espana)
    "Suecia": (8, 22),             # Potter: Brighton, Chelsea
    "Túnez": (14, 8),              # Trabelsi: selecciones/clubes de Tunez
    "Bélgica": (8, 30),            # Garcia: titulo Ligue1 con Lille, Roma, Napoli
    "Egipto": (16, 14),            # Hossam Hassan: leyenda; clubes egipcios
    "Irán": (18, 14),              # Ghalenoei: profundidad en Copa Asia; liga irani
    "Nueva Zelanda": (10, 6),      # Bazeley
    "España": (34, 6),             # De la Fuente: Euro 2024 + Nations League + sub21
    "Cabo Verde": (14, 6),         # Bubista: clasificacion historica al Mundial
    "Arabia Saudita": (8, 14),     # Donis: clubes griegos y sauditas
    "Uruguay": (30, 26),           # Bielsa: oro olimpico 2004; Athletic, Leeds (culto)
    "Francia": (50, 28),           # Deschamps: campeon mundial 2018 + final 2022
    "Senegal": (12, 8),            # Pape Thiaw: interino -> titular
    "Irak": (20, 14),             # Arnold: 2 mundiales con Australia; Sydney FC
    "Noruega": (18, 20),           # Solbakken: Champions con Copenhague
    "Argentina": (48, 4),          # Scaloni: campeon mundial 2022 + 2 Copa America
    "Argelia": (26, 18),           # Petkovic: Euro2020 R16 con Suiza; Lazio
    "Austria": (22, 32),           # Rangnick: arquitecto del proyecto Red Bull
    "Jordania": (16, 8),           # Sellami: final Copa Asia 2023 con Jordania
    "Portugal": (28, 22),          # Roberto Martinez: Belgica #1 FIFA; FA Cup
    "RD Congo": (18, 10),          # Desabre: SF AFCON 2023; Uganda
    "Uzbekistán": (8, 18),         # Cannavaro: campeon mundial 2006 (jugador); China
    "Colombia": (22, 12),          # Lorenzo: final Copa America 2024, larga racha
    "Inglaterra": (12, 40),        # Tuchel: Champions con Chelsea, PSG, Bayern
    "Croacia": (36, 14),           # Dalic: final WC2018 + 3o WC2022
    "Ghana": (28, 18),             # Queiroz: Portugal, Iran x2 WC; asist. Real/Man Utd
    "Panamá": (18, 14),            # Christiansen: clasifico a Panama; APOEL Champions
}

# ===========================================================================
# 2) CLASIFICATORIAS 2026  (PG, PE, PP, GF, GC).  PJ = PG+PE+PP.
#    Anfitriones (Mexico, Canada, EEUU) NO jugaron eliminatoria -> host=True.
#    Estimaciones del ciclo clasificatorio 2026.
# ===========================================================================
HOST = "HOST"
CLASIF = {
    "México": HOST, "Canadá": HOST, "Estados Unidos": HOST,
    "Sudáfrica": (6, 2, 2, 14, 8),
    "Corea del Sur": (7, 3, 0, 20, 6),
    "Chequia": (5, 1, 2, 14, 9),
    "Bosnia y Herzegovina": (5, 2, 1, 14, 8),
    "Catar": (6, 2, 2, 16, 9),
    "Suiza": (5, 3, 0, 15, 6),
    "Brasil": (8, 4, 6, 24, 18),
    "Marruecos": (8, 1, 1, 20, 5),
    "Haití": (5, 2, 3, 12, 10),
    "Escocia": (5, 2, 1, 12, 8),
    "Paraguay": (7, 7, 4, 14, 10),
    "Australia": (6, 3, 1, 16, 8),
    "Turquía": (5, 2, 1, 16, 9),
    "Alemania": (6, 1, 1, 18, 7),
    "Curazao": (4, 4, 2, 12, 9),
    "Costa de Marfil": (6, 2, 2, 16, 8),
    "Ecuador": (7, 4, 7, 18, 14),
    "Países Bajos": (6, 2, 0, 19, 8),
    "Japón": (8, 2, 0, 26, 4),
    "Suecia": (4, 2, 2, 14, 10),
    "Túnez": (6, 3, 1, 12, 2),
    "Bélgica": (5, 3, 0, 20, 9),
    "Egipto": (7, 2, 1, 18, 5),
    "Irán": (8, 2, 0, 18, 4),
    "Nueva Zelanda": (7, 1, 0, 22, 3),
    "España": (7, 1, 0, 21, 4),
    "Cabo Verde": (6, 1, 3, 13, 9),
    "Arabia Saudita": (5, 4, 3, 12, 8),
    "Uruguay": (8, 4, 6, 22, 16),
    "Francia": (7, 1, 0, 18, 5),
    "Senegal": (6, 3, 1, 15, 6),
    "Irak": (5, 4, 3, 14, 10),
    "Noruega": (7, 1, 0, 28, 5),
    "Argentina": (12, 2, 4, 31, 14),
    "Argelia": (7, 1, 2, 18, 8),
    "Austria": (7, 0, 1, 18, 6),
    "Jordania": (6, 3, 1, 15, 7),
    "Portugal": (6, 2, 0, 20, 6),
    "RD Congo": (5, 3, 2, 13, 9),
    "Uzbekistán": (6, 3, 1, 14, 6),
    "Colombia": (8, 4, 6, 22, 18),
    "Inglaterra": (8, 0, 0, 22, 3),
    "Croacia": (7, 1, 0, 20, 5),
    "Ghana": (6, 1, 3, 18, 10),
    "Panamá": (5, 3, 2, 14, 9),
}

# Dificultad de cada confederacion, segun los Mundiales que produce (titulos +
# profundidad historica). UEFA y CONMEBOL concentran TODOS los titulos.
DIFICULTAD_CONF = {
    "UEFA": 1.00, "CONMEBOL": 0.95, "CAF": 0.52,
    "CONCACAF": 0.50, "AFC": 0.48, "OFC": 0.20,
}
# Para anfitriones (sin eliminatoria) se asume un rendimiento por encima del
# promedio en su confederacion (habrian clasificado con comodidad).
HOST_PCT_PTS = 0.70

# ===========================================================================
# 3) JUGADORES EN TOP-5 LIGAS  (conteo sobre un plantel tipico de 26).
#    Proporcion = conteo / 26 (se deriva en features.py).
# ===========================================================================
TOP5 = {
    "México": 6, "Sudáfrica": 4, "Corea del Sur": 8, "Chequia": 9,
    "Canadá": 12, "Bosnia y Herzegovina": 13, "Catar": 0, "Suiza": 17,
    "Brasil": 21, "Marruecos": 19, "Haití": 6, "Escocia": 13,
    "Estados Unidos": 12, "Paraguay": 8, "Australia": 6, "Turquía": 12,
    "Alemania": 24, "Curazao": 7, "Costa de Marfil": 17, "Ecuador": 12,
    "Países Bajos": 22, "Japón": 13, "Suecia": 13, "Túnez": 9,
    "Bélgica": 20, "Egipto": 6, "Irán": 5, "Nueva Zelanda": 4,
    "España": 24, "Cabo Verde": 8, "Arabia Saudita": 0, "Uruguay": 16,
    "Francia": 25, "Senegal": 19, "Irak": 2, "Noruega": 15,
    "Argentina": 20, "Argelia": 17, "Austria": 17, "Jordania": 1,
    "Portugal": 22, "RD Congo": 13, "Uzbekistán": 2, "Colombia": 15,
    "Inglaterra": 26, "Croacia": 19, "Ghana": 15, "Panamá": 3,
}


# ---------------------------------------------------------------------------
def _mapa_pais_fila(ws, col_pais=2, fila_ini=3, fila_fin=50):
    """Devuelve {pais: fila} leyendo la columna de País (1-indexed)."""
    m = {}
    for r in range(fila_ini, fila_fin + 1):
        v = ws.cell(r, col_pais).value
        if v is not None and str(v).strip():
            m[str(v).strip()] = r
    return m


def _slot_a_formula(slot):
    """'1o C' -> formula INDEX/SUMPRODUCT sobre Posiciones; terceros -> None."""
    s = str(slot).strip()
    if "/" in s:                      # slot de mejor tercero -> lo resuelve Python
        return None
    pos = s[0]                        # '1' o '2'
    grupo = s[1:].strip().lstrip("ºoO°").strip()
    if not grupo or not pos.isdigit():
        return None
    # SUMPRODUCT devuelve el nro de fila absoluta; -2 lo lleva a indice del rango.
    return (f'=INDEX(Posiciones!$C$3:$C$50,SUMPRODUCT('
            f'(Posiciones!$A$3:$A$50="{grupo}")*'
            f'(Posiciones!$L$3:$L$50={pos})*'
            f'ROW(Posiciones!$A$3:$A$50))-2)')


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=False)

    # --- 1) DTs: agregar columnas H/I/J con el puntaje (literal) ---
    ws = wb["DTs"]
    ws.cell(2, 8, "Punt. selección (0-50)")
    ws.cell(2, 9, "Punt. clubes (0-50)")
    ws.cell(2, 10, "Puntaje DT (0-100)")
    filas = _mapa_pais_fila(ws)
    falt = []
    for pais, r in filas.items():
        if pais in DT_SCORE:
            sel, clu = DT_SCORE[pais]
            ws.cell(r, 8, sel); ws.cell(r, 9, clu); ws.cell(r, 10, sel + clu)
        else:
            falt.append(pais)
    assert not falt, f"DT_SCORE faltan: {falt}"

    # --- 2) Clasificatorias: D:I crudos + O dificultad + P puntaje ponderado ---
    ws = wb["Clasificatorias"]
    ws.cell(2, 15, "Dificultad conf.")
    ws.cell(2, 16, "Puntaje clasif. ponderado")
    filas = _mapa_pais_fila(ws)
    conf_de = {r: str(ws.cell(r, 3).value).strip() for r in filas.values()}
    falt = []
    for pais, r in filas.items():
        conf = conf_de[r]
        dif = DIFICULTAD_CONF.get(conf, 0.5)
        ws.cell(r, 15, dif)
        reg = CLASIF.get(pais)
        if reg is None:
            falt.append(pais); continue
        if reg == HOST:
            # Anfitrion: sin eliminatoria. Dejar D:I vacios; puntaje = proxy.
            for c in range(4, 10):
                ws.cell(r, c, None)
            punt = round(HOST_PCT_PTS * dif * 100, 1)
        else:
            pg, pe, pp, gf, gc = reg
            pj = pg + pe + pp
            ws.cell(r, 4, pj); ws.cell(r, 5, pg); ws.cell(r, 6, pe)
            ws.cell(r, 7, pp); ws.cell(r, 8, gf); ws.cell(r, 9, gc)
            pct = (3 * pg + pe) / (3 * pj) if pj else 0.0
            punt = round(pct * dif * 100, 1)
        ws.cell(r, 16, punt)
    assert not falt, f"CLASIF faltan: {falt}"

    # --- 3) Predictores_pais: columna I "Jug. en top-5 ligas" (literal) ---
    ws = wb["Predictores_país"]
    filas = _mapa_pais_fila(ws)
    falt = [p for p in filas if p not in TOP5]
    assert not falt, f"TOP5 faltan: {falt}"
    for pais, r in filas.items():
        ws.cell(r, 9, TOP5[pais])

    # --- 4) Eliminatorias: slots 1o/2o -> formula contra Posiciones ---
    ws = wb["Eliminatorias"]
    # Asegurar columnas Pen 1 / Pen 2 (tanda de penales) al final, idempotente.
    # Header en fila 2. Sólo se usan para desempatar un KO igualado en los 90'.
    hdr = {ws.cell(2, c).value: c for c in range(1, ws.max_column + 1)}
    if "Pen 1" not in hdr:
        c1 = ws.max_column + 1
        ws.cell(2, c1, "Pen 1")
        ws.cell(2, c1 + 1, "Pen 2")
    n_form = 0
    for r in range(3, ws.max_row + 1):
        slot1 = ws.cell(r, 9).value   # col I
        slot2 = ws.cell(r, 10).value  # col J
        if slot1:
            f = _slot_a_formula(slot1)
            if f:
                ws.cell(r, 3, f); n_form += 1     # col C = Equipo 1
        if slot2:
            f = _slot_a_formula(slot2)
            if f:
                ws.cell(r, 6, f); n_form += 1     # col F = Equipo 2

    # --- 5) Leeme: actualizar nota de corte ---
    ws = wb["Léeme"]
    ws.cell(2, 3, ("Cargados 54 resultados de fase de grupos (marcador exacto): "
                   "fechas 1 y 2 completas de los 12 grupos + parte de la fecha 3. "
                   "El resto se simula. Variables curadas agregadas: puntaje de DT, "
                   "clasificatorias ponderadas por confederacion y % de jugadores "
                   "en top-5 ligas (estimaciones ~early-2026). "
                   "Fuente resultados: Yahoo Sports / Wikipedia por grupo."))

    wb.save(XLSX)
    print(f"OK. Excel enriquecido. Formulas de slots escritas: {n_form}.")
    print("Hojas tocadas: DTs, Clasificatorias, Predictores_país, Eliminatorias, Léeme.")


if __name__ == "__main__":
    main()
